# -*- coding: utf-8 -*-
"""Parse specialty groups with dates → journals_structured.xlsx."""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import AS_OF_LABEL, STRUCTURED_XLSX, pdf_path_for_parsing
from extract_journals import ISSN_RE, extract_full_text, normalize_issn, parse_entry, split_entries

DATE_LINE = re.compile(r"^([сС]|по)\s+(\d{2}\.\d{2}\.\d{4})$")
SPEC_START = re.compile(
    r"^("
    r"(?P<diss>\d{2}\.\d{2}\.\d{2})\s*[–—\-]\s*"
    r"|(?P<vak>\d{1,2}\.\d{1,2}\.\d{1,2})\.?\s*"
    r")"
)
BRANCH_RE = re.compile(r"\(([^)]+)\)\s*$")
# Split when a new specialty code appears mid-line (after ISSN / rename notes)
EMBEDDED_SPEC = re.compile(
    r"(?=(?:\d{1,2}\.\d{1,2}\.\d{1,2}\.?\s*)|(?:\d{2}\.\d{2}\.\d{2}\s*[–—\-]\s))"
)
ISSN_INLINE = re.compile(r"[,;\s]*ISSN\s+[\dXxХх\-–—\s,\)»«\"]+", re.IGNORECASE)
BRANCH_CONT = re.compile(r"^\([^)]+\)\s*,?\s*$")
ENTRY_START = re.compile(r"^(\d{1,4})\.\s+(?!\d)")


@dataclass
class SpecRecord:
    code: str
    code_type: str
    title: str
    branch: str = ""
    date_from: str = ""
    date_to: str = ""
    dates_raw: str = ""
    group_index: int = 0


@dataclass
class JournalSpecs:
    num: int
    name: str
    issn: str
    journal_dates: str
    specs: list[SpecRecord] = field(default_factory=list)
    parse_notes: str = ""


def parse_dates_list(date_strs: list[str]) -> tuple[str, str, str]:
    froms, tos = [], []
    for d in date_strs:
        kind, val = d.split(" ", 1)
        if kind in ("с", "С"):
            froms.append(val)
        elif kind == "по":
            tos.append(val)
    return froms[0] if froms else "", tos[-1] if tos else "", "; ".join(date_strs)


def extract_branch(title: str) -> tuple[str, str]:
    title = title.strip(" ,")
    m = BRANCH_RE.search(title)
    if not m:
        return clean_spec_title(title), ""
    core = clean_spec_title(title[: m.start()].strip(" ,"))
    return core, m.group(1).strip()


def clean_spec_title(title: str) -> str:
    """Remove ISSN / rename-note debris often injected by PDF line breaks."""
    t = ISSN_INLINE.sub("", title)
    t = re.sub(r"\s+", " ", t).strip(" ,;")
    return t


def is_issn_junk_line(line: str) -> bool:
    """Lines that are only old ISSN / title fragments, not specialties."""
    if SPEC_START.match(line):
        return False
    if not re.search(r"ISSN", line, re.IGNORECASE):
        return False
    rest = re.sub(r"ISSN\s*", "", line, flags=re.IGNORECASE)
    rest = re.sub(r"[\dXxХх\-–—\s,\)»«\"«]+", "", rest)
    return len(rest) < 2


def junk_before_next_spec(lines: list[str], start: int) -> bool:
    """True when ISSN/rename debris sits between a date row and the next specialty."""
    saw_junk = False
    for j in range(start, len(lines)):
        line = lines[j].strip()
        if not line:
            continue
        for part in expand_spec_lines([line]):
            if is_issn_junk_line(part):
                saw_junk = True
                continue
            if SPEC_START.match(part):
                return saw_junk
    return False


def peek_next_is_spec(lines: list[str], start: int) -> bool:
    for j in range(start, len(lines)):
        line = lines[j].strip()
        if not line:
            continue
        for part in expand_spec_lines([line]):
            if is_issn_junk_line(part):
                continue
            return bool(SPEC_START.match(part))
    return False


def peek_raw_has_spec_ahead(raw_lines: list[str], start: int) -> bool:
    """Specialty after rename/ISSN notes that PDF did not expand yet."""
    for j in range(start, len(raw_lines)):
        line = raw_lines[j].strip()
        if not line:
            continue
        if peek_next_is_spec(expand_spec_lines([line]), 0):
            return True
        if SPEC_START.match(line) or EMBEDDED_SPEC.search(line):
            return True
    return False


def raw_junk_before_next_spec(raw_lines: list[str], start: int) -> bool:
    """Detect rename/ISSN notes on raw PDF lines (before line splitting)."""
    for j in range(start, len(raw_lines)):
        line = raw_lines[j].strip()
        if not line:
            continue
        if re.search(r"ISSN\s+\d", line, re.IGNORECASE):
            return True
        if SPEC_START.match(line) or EMBEDDED_SPEC.search(line):
            return False
    return False


def is_split_row_date(
    current_specs: list[dict],
    lines: list[str],
    after_i: int,
    raw_lines: list[str],
    raw_after: int,
) -> bool:
    """
  Date cell spans pages within one table row: specs, then «с …», then ISSN junk,
  then more specs on the next page (e.g. №466: 5.3.1/5.3.3 + с 01.02.2022 + 5.3.4).
    """
    if not current_specs:
        return False
    has_junk = junk_before_next_spec(lines, after_i) or raw_junk_before_next_spec(
        raw_lines, raw_after
    )
    if not has_junk:
        return False
    return peek_next_is_spec(lines, after_i) or peek_raw_has_spec_ahead(
        raw_lines, raw_after
    )


def dates_apply_to_following(
    batch: list[str],
    current_specs: list[dict],
    lines: list[str],
    after_i: int,
    raw_lines: list[str],
    raw_after: int,
) -> bool:
    """
    Date on the previous page belongs to the next specialty only (no specs in the
    current row buffer yet). Not the split-row case handled by is_split_row_date.
    """
    if current_specs or not peek_next_is_spec(lines, after_i):
        return False
    return junk_before_next_spec(lines, after_i) or raw_junk_before_next_spec(
        raw_lines, raw_after
    )


def has_branch_in_title(title: str) -> bool:
    t = title.strip()
    return bool(BRANCH_RE.search(t)) or bool(re.search(r"\([^)]+\)", t))


def is_branch_continuation_line(line: str) -> bool:
    """Title fragment continuing previous specialty cell (often after page break)."""
    s = line.strip()
    if not s or SPEC_START.match(s) or DATE_LINE.match(s):
        return False
    return bool(BRANCH_CONT.match(s)) or (
        s.startswith("(") and ")" in s and not EMBEDDED_SPEC.search(s)
    )


def page_leading_dates(page_lines: list[str]) -> list[str] | None:
    """First table row on a page starts with date(s) in the last column."""
    batch: list[str] = []
    for line in page_lines:
        s = line.strip()
        if not s:
            if batch:
                break
            continue
        if ENTRY_START.match(s):
            return None
        dm = DATE_LINE.match(s)
        if dm:
            kind = "с" if dm.group(1) in ("с", "С") else dm.group(1)
            batch.append(f"{kind} {dm.group(2)}")
            continue
        if batch:
            break
        return None
    return batch or None


def page_start_has_date(page_lines: list[str]) -> bool:
    return page_leading_dates(page_lines) is not None


def apply_page_start(
    page_lines: list[str],
    *,
    current_specs: list[dict],
    current: dict | None,
    segment_dates: list[str],
    split_row_continue: bool,
    pending_leading_dates: list[str] | None,
    ended_mid_row: bool,
) -> tuple[dict | None, list[str], bool, list[str] | None, int]:
    """
    Apply table rules at a page boundary. Returns
    (current, segment_dates, split_row_continue, pending_leading_dates, skip_lines).
    """
    skip = 0
    if not page_lines:
        return current, segment_dates, split_row_continue, pending_leading_dates, skip

    leading = page_leading_dates(page_lines)
    if leading is not None:
        consumed = 0
        for line in page_lines:
            s = line.strip()
            if not s:
                if consumed:
                    break
                continue
            if DATE_LINE.match(s):
                consumed += 1
                continue
            break
        return current, segment_dates, False, leading, consumed

    if not ended_mid_row:
        return current, segment_dates, False, pending_leading_dates, skip

    # No date in the first row of the date column → row continues from previous page.
    split_row_continue = True
    idx = 0
    while idx < len(page_lines) and not page_lines[idx].strip():
        idx += 1
    if idx >= len(page_lines):
        return current, segment_dates, split_row_continue, pending_leading_dates, skip

    first = page_lines[idx].strip()
    if is_branch_continuation_line(first) and current_specs:
        target = current if current is not None else current_specs[-1]
        target["title"] = (target["title"] + " " + first).strip()
        if current is None:
            current = target
        return current, segment_dates, split_row_continue, pending_leading_dates, idx + 1

    return current, segment_dates, split_row_continue, pending_leading_dates, skip


def date_ahead_differs_from_segment(
    raw_lines: list[str], start: int, segment_dates: list[str]
) -> bool:
    for j in range(start, len(raw_lines)):
        s = raw_lines[j].strip()
        if not s:
            continue
        dm = DATE_LINE.match(s)
        if dm:
            kind = "с" if dm.group(1) in ("с", "С") else dm.group(1)
            upcoming = f"{kind} {dm.group(2)}"
            return upcoming not in segment_dates
    return False


def spec_has_trailing_date_ahead(raw_lines: list[str], start: int) -> bool:
    """Next date line appears before the next specialty (new row with own dates)."""
    for j in range(start, len(raw_lines)):
        line = raw_lines[j].strip()
        if not line:
            continue
        if DATE_LINE.match(line):
            return True
        for part in expand_spec_lines([line]):
            if is_issn_junk_line(part):
                continue
            if SPEC_START.match(part):
                return False
    return False


def expand_spec_lines(lines: list[str]) -> list[str]:
    """One PDF line may contain 'ISSN …) 5.3.4. Title' — split before tokenizing."""
    out: list[str] = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        embedded = EMBEDDED_SPEC.search(line)
        if embedded and embedded.start() > 0:
            parts = [p.strip() for p in EMBEDDED_SPEC.split(line) if p.strip()]
        else:
            parts = [line]
        for part in parts:
            if not is_issn_junk_line(part):
                out.append(part)
    return out


def tokenize_spec_region(lines: list[str]) -> list[dict]:
    tokens: list[dict] = []
    buf: list[str] = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        dm = DATE_LINE.match(line)
        if dm:
            if buf:
                tokens.append({"type": "text", "text": " ".join(buf)})
                buf = []
            kind = "с" if dm.group(1) in ("с", "С") else dm.group(1)
            tokens.append({"type": "date", "kind": kind, "value": dm.group(2)})
            continue
        sm = SPEC_START.match(line)
        if sm:
            if buf:
                tokens.append({"type": "text", "text": " ".join(buf)})
                buf = []
            code = sm.group("vak") or sm.group("diss")
            tokens.append(
                {
                    "type": "spec_start",
                    "code": code + ("" if sm.group("diss") else "."),
                    "code_type": "vak" if sm.group("vak") else "diss",
                    "rest": line[sm.end() :].strip(),
                }
            )
            continue
        buf.append(line)
    if buf:
        tokens.append({"type": "text", "text": " ".join(buf)})
    return tokens


def parse_specs_segmented(after_issn: str) -> tuple[list[SpecRecord], str]:
    page_chunks = [
        p for p in re.split(r"\s*<<<PDF_PAGE>>>\s*", after_issn) if p.strip()
    ]
    raw_pages: list[list[str]] = []
    for chunk in page_chunks:
        raw_pages.append([l.strip() for l in chunk.splitlines() if l.strip()])

    segments: list[tuple[list[dict], list[str]]] = []
    current_specs: list[dict] = []
    segment_dates: list[str] = []
    current: dict | None = None
    notes: list[str] = []
    split_row_active = False
    split_row_continue = False
    pending_leading_dates: list[str] | None = None
    specs_on_page_since_break = 0

    def close_segment():
        nonlocal current_specs, segment_dates, split_row_continue
        if current_specs:
            segments.append((current_specs, list(segment_dates)))
        current_specs = []
        segment_dates = []
        split_row_continue = False

    def page_end_split_row(page_i: int, line_i: int) -> bool:
        if not current_specs or page_i + 1 >= len(raw_pages):
            return False
        next_page = raw_pages[page_i + 1]
        if page_start_has_date(next_page):
            return False
        rest = raw_page[line_i:]
        if rest and not all(not ln.strip() or DATE_LINE.match(ln.strip()) for ln in rest):
            return False
        return peek_raw_has_spec_ahead(next_page, 0) or any(
            is_branch_continuation_line(ln)
            for ln in next_page
            if ln.strip()
        )

    def handle_date_batch(
        batch: list[str], raw_page: list[str], line_i: int, page_i: int
    ) -> None:
        nonlocal segment_dates, split_row_active, split_row_continue, pending_leading_dates, current
        tail = expand_spec_lines(raw_page[line_i:])
        tail_raw = raw_page[line_i:]
        if current is not None and current_specs and not has_branch_in_title(
            current["title"]
        ):
            segment_dates = batch
            split_row_active = True
            split_row_continue = True
        elif current is not None and current_specs and has_branch_in_title(
            current["title"]
        ):
            segment_dates = batch
            current = None
            split_row_active = False
        elif is_split_row_date(current_specs, tail, 0, tail_raw, 0):
            segment_dates = batch
            split_row_active = True
            split_row_continue = True
        elif dates_apply_to_following(batch, current_specs, tail, 0, tail_raw, 0):
            close_segment()
            pending_leading_dates = batch
            split_row_active = False
        else:
            segment_dates = batch
            split_row_active = False
        if page_end_split_row(page_i, line_i):
            split_row_active = True
            split_row_continue = True

    for page_i, raw_page in enumerate(raw_pages):
        skip = 0
        if page_i > 0:
            ended_mid_row = current is not None or split_row_active
            specs_on_page_since_break = 0
            (
                current,
                segment_dates,
                split_row_continue,
                pending_leading_dates,
                skip,
            ) = apply_page_start(
                raw_page,
                current_specs=current_specs,
                current=current,
                segment_dates=segment_dates,
                split_row_continue=split_row_continue,
                pending_leading_dates=pending_leading_dates,
                ended_mid_row=ended_mid_row,
            )
            if split_row_continue and segment_dates:
                split_row_active = True

        raw_i = skip
        while raw_i < len(raw_page):
            raw_line = raw_page[raw_i]
            parts = expand_spec_lines([raw_line]) if raw_line.strip() else []
            if not parts:
                raw_i += 1
                continue

            if DATE_LINE.match(raw_line.strip()):
                batch: list[str] = []
                batch_end = raw_i
                while batch_end < len(raw_page) and DATE_LINE.match(
                    raw_page[batch_end].strip()
                ):
                    dm2 = DATE_LINE.match(raw_page[batch_end].strip())
                    kind = "с" if dm2.group(1) in ("с", "С") else dm2.group(1)
                    batch.append(f"{kind} {dm2.group(2)}")
                    batch_end += 1
                handle_date_batch(batch, raw_page, batch_end, page_i)
                raw_i = batch_end
                continue

            for line in parts:
                sm = SPEC_START.match(line)
                if sm:
                    if pending_leading_dates is not None:
                        if current_specs and segment_dates:
                            close_segment()
                        segment_dates = list(pending_leading_dates)
                        pending_leading_dates = None
                        split_row_active = False
                        split_row_continue = False
                    elif current_specs and segment_dates:
                        if split_row_active or split_row_continue:
                            ahead_new = date_ahead_differs_from_segment(
                                raw_page, raw_i + 1, segment_dates
                            )
                            close_before = False
                            if ahead_new and specs_on_page_since_break >= 2:
                                close_before = True
                            elif (
                                ahead_new
                                and specs_on_page_since_break >= 1
                                and spec_has_trailing_date_ahead(
                                    raw_page, raw_i + 1
                                )
                            ):
                                close_before = True
                            if close_before:
                                close_segment()
                                split_row_active = False
                                split_row_continue = False
                                specs_on_page_since_break = 0
                        else:
                            close_segment()
                    code = sm.group("vak") or sm.group("diss")
                    vak = sm.group("vak")
                    current = {
                        "code": code + ("." if vak and not code.endswith(".") else ""),
                        "code_type": "vak" if vak else "diss",
                        "title": line[sm.end() :].strip(),
                    }
                    current_specs.append(current)
                    specs_on_page_since_break += 1
                    continue
                if current is not None:
                    current["title"] = (current["title"] + " " + line).strip()
                elif not is_issn_junk_line(line):
                    notes.append(f"orphan_text:{line[:40]}")
            raw_i += 1

    if current_specs:
        close_segment()

    records: list[SpecRecord] = []
    for gi, (specs, dates) in enumerate(segments):
        d_from, d_to, d_raw = parse_dates_list(dates)
        for s in specs:
            title, branch = extract_branch(s["title"])
            inline = [
                f"{'с' if k in ('с', 'С') else k} {v}"
                for k, v in re.findall(r"\b([сС]|по)\s+(\d{2}\.\d{2}\.\d{4})\b", title)
            ]
            if inline:
                title = re.sub(r"\s*([сС]|по)\s+\d{2}\.\d{2}\.\d{4}\s*", " ", title).strip()
                id_from, id_to, id_raw = parse_dates_list(inline)
                d_from, d_to = d_from or id_from, d_to or id_to
                d_raw = "; ".join(x for x in (d_raw, id_raw) if x)
            records.append(
                SpecRecord(
                    code=s["code"],
                    code_type=s["code_type"],
                    title=title,
                    branch=branch,
                    date_from=d_from,
                    date_to=d_to,
                    dates_raw=d_raw,
                    group_index=gi,
                )
            )
    return records, "; ".join(notes)


def fill_missing_spec_dates(specs: list[SpecRecord], journal_dates: str) -> None:
    """Journal-level fallback only when no specialty received parsed dates."""
    if any(s.dates_raw for s in specs):
        return
    j_from, j_to, j_raw = parse_dates_list(
        [d.strip() for d in journal_dates.split(";") if d.strip()]
    )
    if not j_from and not j_to:
        return
    for s in specs:
        s.date_from, s.date_to, s.dates_raw = j_from, j_to, j_raw


def find_spec_region_start(block: str) -> int:
    matches = list(ISSN_RE.finditer(block))
    if not matches:
        return 0
    for m in reversed(matches):
        tail = block[m.end() : m.end() + 80]
        if SPEC_START.search(tail.strip()) or re.search(
            r"^\d{1,2}\.\d{1,2}\.\d{1,2}\.", tail.strip()
        ):
            return m.end()
    return matches[-1].end()


def parse_journal(num: int, block: str, journal_row: dict) -> JournalSpecs:
    matches = list(ISSN_RE.finditer(block))
    start = find_spec_region_start(block)
    after = block[start:]
    issn = journal_row["issn"]
    if matches:
        for m in reversed(matches):
            if m.end() <= start + 2:
                issn = normalize_issn(m)
                break
    specs, notes = parse_specs_segmented(after)
    fill_missing_spec_dates(specs, journal_row.get("dates", ""))
    return JournalSpecs(
        num=num,
        name=journal_row["name"],
        issn=issn,
        journal_dates=journal_row.get("dates", ""),
        specs=specs,
        parse_notes=notes,
    )


def parse_all(pdf_path: Path) -> list[JournalSpecs]:
    blocks = dict(split_entries(extract_full_text(pdf_path)))
    return [
        parse_journal(num, blocks[num], parse_entry(num, blocks[num]))
        for num in sorted(blocks)
    ]


def write_xlsx(journals: list[JournalSpecs], out_path: Path) -> dict:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    catalog: dict[tuple, dict] = {}
    mapping_rows: list[dict] = []
    stats: dict = defaultdict(int)

    for j in journals:
        stats["journals"] += 1
        for s in j.specs:
            stats["links"] += 1
            if s.dates_raw:
                stats["links_with_dates"] += 1
            k = (s.code_type, s.code, s.title, s.branch)
            if k not in catalog:
                catalog[k] = {
                    "spec_id": len(catalog) + 1,
                    "code_type": s.code_type,
                    "code": s.code,
                    "title": s.title,
                    "branch": s.branch,
                    "journal_count": 0,
                }
            catalog[k]["journal_count"] += 1
            mapping_rows.append(
                {
                    "journal_num": j.num,
                    "journal_name": j.name,
                    "issn": j.issn,
                    "spec_id": catalog[k]["spec_id"],
                    "code_type": s.code_type,
                    "code": s.code,
                    "title": s.title,
                    "branch": s.branch,
                    "group_index": s.group_index,
                    "date_from": s.date_from,
                    "date_to": s.date_to,
                    "dates_raw": s.dates_raw,
                    "journal_dates_all": j.journal_dates,
                }
            )

    wb = Workbook()

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_j = wb.active
    ws_j.title = "Journals"
    ws_j.append(
        ["№", "Наименование", "ISSN", "Даты (журнал)", "Число специальностей", "Примечания"]
    )
    for j in journals:
        ws_j.append([j.num, j.name, j.issn, j.journal_dates, len(j.specs), j.parse_notes])
    style_header(ws_j)
    autosize(ws_j, [6, 42, 14, 28, 10, 20])
    ws_j.freeze_panes = "A2"

    ws_s = wb.create_sheet("Specialties")
    ws_s.append(["spec_id", "code_type", "code", "title", "branch", "journal_count"])
    for row in sorted(catalog.values(), key=lambda x: (x["code_type"], x["code"])):
        ws_s.append(
            [row["spec_id"], row["code_type"], row["code"], row["title"], row["branch"], row["journal_count"]]
        )
    style_header(ws_s)
    autosize(ws_s, [8, 10, 12, 55, 28, 12])

    ws_m = wb.create_sheet("Journal_Spec_Map")
    ws_m.append(
        [
            "journal_num", "journal_name", "issn", "spec_id", "code_type", "code",
            "title", "branch", "group_index", "date_from", "date_to", "dates_raw",
            "journal_dates_all",
        ]
    )
    for r in mapping_rows:
        ws_m.append(list(r.values()))
    style_header(ws_m)
    autosize(ws_m, [8, 42, 14, 8, 10, 12, 45, 25, 8, 12, 12, 28, 28])
    ws_m.freeze_panes = "A2"

    ws_p = wb.create_sheet("Parse_Summary")
    ws_p.append(["metric", "value"])
    for k, v in [
        ("journals", stats["journals"]),
        ("links", stats["links"]),
        ("links_with_dates", stats["links_with_dates"]),
        ("unique_specialties", len(catalog)),
    ]:
        ws_p.append([k, v])
    style_header(ws_p)
    wb.save(out_path)
    return dict(stats) | {"unique_specialties": len(catalog)}


def main() -> int:
    p = argparse.ArgumentParser(description=f"Parse specialties + dates ({AS_OF_LABEL})")
    p.add_argument("--pdf", type=Path, default=None, help="Input PDF (default: SOURCE_URL.txt)")
    p.add_argument("--output", type=Path, default=STRUCTURED_XLSX)
    args = p.parse_args()

    try:
        pdf = args.pdf or pdf_path_for_parsing()
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1

    print(f"Parsing {pdf}…")
    journals = parse_all(pdf)
    links = sum(len(j.specs) for j in journals)
    dated = sum(1 for j in journals for s in j.specs if s.dates_raw)
    print(f"  {len(journals)} journals, {links} links ({dated} with dates)")
    stats = write_xlsx(journals, args.output)
    print(f"Wrote {args.output}")
    print(stats)
    return 0


if __name__ == "__main__":
    sys.exit(main())
