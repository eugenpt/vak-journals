# -*- coding: utf-8 -*-
"""Extract the journal table from the VAK list PDF → journals.xlsx."""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from config import AS_OF_LABEL, JOURNALS_XLSX, PDF_PATH, pdf_path_for_parsing

ENTRY_START = re.compile(r"^(\d{1,4})\.\s+(?!\d)")
ISSN_RE = re.compile(r"\b(\d{4})[\s\-–—]+(\d{3}[\dXxХх])\b")
DATE_RE = re.compile(r"\b((?:с|по)\s+\d{2}\.\d{2}\.\d{4})\b", re.IGNORECASE)
SPEC_LINE = re.compile(
    r"^("
    r"\d{1,2}\.\d{1,2}\.\d{1,2}\."
    r"|\d{2}\.\d{2}\.\d{2}\s*[–—\-]"
    r"|\d{2}\.\d{2}\.\d{2}\s"
    r")",
    re.MULTILINE,
)
HEADER_MARKERS = (
    "ПЕРЕЧЕНЬ",
    "рецензируемых научных",
    "п/п Наименование",
    "Дата включения",
)


# Form feed is stripped by str.splitlines(); use a line marker instead.
PAGE_BREAK = "\n<<<PDF_PAGE>>>\n"


def extract_full_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    return PAGE_BREAK.join(page.extract_text() or "" for page in reader.pages)


def is_header_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    return any(m in s for m in HEADER_MARKERS)


def split_entries(text: str) -> list[tuple[int, str]]:
    lines = text.splitlines()
    entries: list[tuple[int, str]] = []
    current_num: int | None = None
    current_lines: list[str] = []

    for line in lines:
        if is_header_line(line):
            continue
        m = ENTRY_START.match(line)
        if m:
            num = int(m.group(1))
            if current_num is not None:
                entries.append((current_num, "\n".join(current_lines)))
            current_num = num
            current_lines = [line]
        elif current_num is not None:
            current_lines.append(line)

    if current_num is not None:
        entries.append((current_num, "\n".join(current_lines)))
    return entries


def normalize_issn(match: re.Match) -> str:
    suffix = match.group(2).replace("Х", "X").replace("х", "x")
    return f"{match.group(1)}-{suffix}"


def parse_entry(num: int, block: str) -> dict:
    issn_match = ISSN_RE.search(block)
    issn = normalize_issn(issn_match) if issn_match else ""

    seen: set[str] = set()
    ordered_dates: list[str] = []
    for d in (m.group(1) for m in DATE_RE.finditer(block)):
        if d not in seen:
            seen.add(d)
            ordered_dates.append(d)

    if issn_match:
        before_issn = block[: issn_match.start()]
        after_issn = block[issn_match.end() :]
    else:
        before_issn = block
        after_issn = ""

    title_lines = []
    for i, line in enumerate(before_issn.splitlines()):
        if i == 0:
            line = ENTRY_START.sub("", line).strip()
        line = line.strip()
        if line:
            title_lines.append(line)
    name = re.sub(r"\s+", " ", " ".join(title_lines)).strip()

    spec_parts = []
    if after_issn:
        for line in after_issn.splitlines():
            line = line.strip()
            if not line:
                continue
            if DATE_RE.search(line) and not SPEC_LINE.match(line):
                remaining = DATE_RE.sub("", line).strip()
                if remaining and SPEC_LINE.match(remaining):
                    spec_parts.append(remaining)
                continue
            cleaned = DATE_RE.sub("", line).strip()
            if cleaned:
                spec_parts.append(cleaned)

    specialties = "\n".join(spec_parts)
    specialties = re.sub(r"[ \t]+", " ", specialties)
    specialties = re.sub(r"\n{3,}", "\n\n", specialties).strip()

    for d in (m.group(1) for m in DATE_RE.finditer(specialties)):
        if d not in seen:
            seen.add(d)
            ordered_dates.append(d)
    specialties = DATE_RE.sub("", specialties)
    specialties = re.sub(r"\s{2,}", " ", specialties)
    specialties = re.sub(r"\n{2,}", "\n", specialties).strip(" \n,")

    return {
        "num": num,
        "name": name,
        "issn": issn,
        "specialties": specialties,
        "dates": "; ".join(ordered_dates),
    }


def parse_all(pdf_path: Path) -> list[dict]:
    rows = [parse_entry(n, b) for n, b in split_entries(extract_full_text(pdf_path))]
    rows.sort(key=lambda r: r["num"])
    return rows


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Перечень"
    headers = [
        "№",
        "Наименование издания",
        "ISSN",
        "Научные специальности и отрасли науки",
        "Даты включения / исключения",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row["num"], row["name"], row["issn"], row["specialties"], row["dates"]])
    for i, w in enumerate([6, 45, 14, 70, 28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description=f"Extract journals table ({AS_OF_LABEL})")
    p.add_argument(
        "--pdf",
        type=Path,
        default=None,
        help="Input PDF (default: SOURCE_URL.txt — URL cache or local path)",
    )
    p.add_argument("--output", type=Path, default=JOURNALS_XLSX, help="Output XLSX path")
    args = p.parse_args()

    try:
        pdf = args.pdf or pdf_path_for_parsing()
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1

    print(f"Reading {pdf} ({AS_OF_LABEL})…")
    rows = parse_all(pdf)
    print(f"  {len(rows)} journals, {sum(1 for r in rows if not r['issn'])} without ISSN")
    print(f"Writing {args.output}…")
    write_xlsx(rows, args.output)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
